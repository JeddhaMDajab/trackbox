# ======================= IMPORTS =======================
import os
import shutil
from datetime import datetime, timedelta
from typing import Optional, List, Dict
import jwt
from dotenv import load_dotenv
from authlib.integrations.starlette_client import OAuth
from starlette.middleware.sessions import SessionMiddleware

load_dotenv()

def get_ph_time():
    """Returns the current time in Philippine Time (UTC+8)."""
    return datetime.utcnow() + timedelta(hours=8)

from passlib.hash import bcrypt
from fastapi import (
    FastAPI, Request, Form, UploadFile, File,
    Depends, HTTPException, status, WebSocket, WebSocketDisconnect
)
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse, FileResponse, StreamingResponse
import httpx
import io
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

# ======================= REAL-TIME MANAGER =======================
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, WebSocket] = {}

    async def connect(self, username: str, websocket: WebSocket):
        await websocket.accept()
        self.active_connections[username] = websocket

    def disconnect(self, username: str):
        if username in self.active_connections:
            del self.active_connections[username]

    async def send_personal_message(self, message: dict, username: str):
        if username in self.active_connections:
            websocket = self.active_connections[username]
            await websocket.send_json(message)

manager = ConnectionManager()

from sqlalchemy import (
    create_engine, Column, Integer, String, Text, DateTime, Boolean, func, extract
)
from sqlalchemy.orm import sessionmaker, declarative_base, Session

from PIL import Image
import numpy as np
from skimage.metrics import structural_similarity as ssim


# ======================= CONFIG =======================
SECRET_KEY = os.getenv("SECRET_KEY", "change_this_in_production")
ALGORITHM = os.getenv("ALGORITHM", "HS256")
ACCESS_TOKEN_EXPIRE_MINUTES = int(os.getenv("ACCESS_TOKEN_EXPIRE_MINUTES", 1440))

# OAuth Config
oauth = OAuth()
oauth.register(
    name='google',
    client_id=os.getenv("GOOGLE_CLIENT_ID"),
    client_secret=os.getenv("GOOGLE_CLIENT_SECRET"),
    server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
    client_kwargs={
        'scope': 'openid email profile'
    }
)

# --- SMTP CONFIG (FOR REAL EMAILS) ---
# Note: For Gmail, use an "App Password" if 2FA is enabled.
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SMTP_USER = "jeddha.dajab@evsu.edu.ph" # Replace with your EVSU or personal email
SMTP_PASS = "zrof glry shbn fceq"    # Replace with your App Password
MAIL_FROM = "jeddha.dajab@evsu.edu.ph"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    DATABASE_URL = f"sqlite:///{os.path.join(BASE_DIR, 'trackbox.db')}"
elif DATABASE_URL.startswith("mysql://"):
    DATABASE_URL = DATABASE_URL.replace("mysql://", "mysql+pymysql://", 1)
elif DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

UPLOAD_FOLDER = "uploads"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs("static", exist_ok=True)
os.makedirs("templates", exist_ok=True)
os.makedirs("temp_uploads", exist_ok=True)

# ======================= DATABASE =======================
if DATABASE_URL.startswith("sqlite"):
    engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
else:
    engine = create_engine(DATABASE_URL, pool_pre_ping=True)
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()


class Building(Base):
    __tablename__ = "buildings"
    id = Column(Integer, primary_key=True)
    name = Column(String, unique=True, index=True)

class User(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True)
    username = Column(String, unique=True)
    first_name = Column(String)
    middle_name = Column(String, nullable=True)
    last_name = Column(String)
    email = Column(String, unique=True, index=True)
    hashed_password = Column(String)
    role = Column(String, default="student") # student, admin, guard, coordinator
    assigned_building = Column(String, nullable=True) # For US-03 Coordinator/Guard filtering
    is_active = Column(Boolean, default=True) # For US-04 Deactivation
    points = Column(Integer, default=0)
    profile_image = Column(String, nullable=True)
    reset_token = Column(String, nullable=True) # For forgot password logic

class LostItem(Base):
    __tablename__ = "lost_items"
    id = Column(Integer, primary_key=True, index=True)
    reporter = Column(String)
    type = Column(String)
    category = Column(String)
    item_name = Column(String)
    description = Column(String)
    last_seen = Column(String)
    building = Column(String, default="Main Building") # For US-03 filtering
    image = Column(String, nullable=True)
    status = Column(String, default="Pending")
    matched_with = Column(Integer, nullable=True) # ID of the matching item
    security_question = Column(String, nullable=True)
    security_answer = Column(String, nullable=True)
    is_archived = Column(Boolean, default=False)
    is_owner_verified = Column(Boolean, default=False)
    verified_at = Column(DateTime, nullable=True)
    handed_over_by = Column(String, nullable=True)
    handover_method = Column(String, nullable=True)
    claimed_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=get_ph_time)



class FoundItem(Base):
    __tablename__ = "found_items"
    id = Column(Integer, primary_key=True)
    reporter = Column(String)
    category = Column(String)
    item_name = Column(String)
    description = Column(Text)
    found_in = Column(String)
    building = Column(String, default="Main Building") # For US-03 filtering
    image = Column(String)
    status = Column(String, default="Pending")
    is_claimed = Column(Boolean, default=False)
    is_verified = Column(Boolean, default=False)
    security_question = Column(String, nullable=True)
    security_answer = Column(String, nullable=True)
    is_archived = Column(Boolean, default=False)
    handed_over_by = Column(String, nullable=True) # Usually 'Self' or 'Admin'
    handed_over_to = Column(String, nullable=True) # The owner
    handover_method = Column(String, nullable=True)
    claimed_at = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=get_ph_time)

class Notification(Base):
    __tablename__ = "notifications"
    id = Column(Integer, primary_key=True)
    recipient = Column(String)
    message = Column(String)
    is_read = Column(Boolean, default=False)
    created_at = Column(DateTime, default=get_ph_time)

class Message(Base):
    __tablename__ = "messages"
    id = Column(Integer, primary_key=True)
    sender = Column(String)
    receiver = Column(String)
    content = Column(String)
    item_id = Column(Integer)
    created_at = Column(DateTime, default=datetime.utcnow)

DB_PATH = "database.db"

def get_all_lost_items(db: Session):
    """
    Return all lost and found items as a list of dicts for student search page.
    """
    lost_items = db.query(LostItem).order_by(LostItem.created_at.desc()).all()
    found_items = db.query(FoundItem).order_by(FoundItem.created_at.desc()).all()

    items = []

    for item in lost_items:
        items.append({
            "id": item.id,
            "title": item.item_name,
            "type": "Lost",
            "description": item.description,
            "location": item.last_seen,
            "image": item.image,
            "created_at": item.created_at.strftime('%Y-%m-%d %H:%M:%S')
        })

    for item in found_items:
        items.append({
            "id": item.id,
            "title": item.item_name,
            "type": "Found",
            "description": item.description,
            "location": item.found_in,
            "image": item.image,
            "created_at": item.created_at.strftime('%Y-%m-%d %H:%M:%S')
        })

    # sort by created_at descending
    items.sort(key=lambda x: x["created_at"], reverse=True)
    return items



Base.metadata.create_all(bind=engine)

# ======================= APP SETUP =======================
app = FastAPI()

# Middleware to handle HTTPS proxy redirection (e.g. Railway TLS termination)
@app.middleware("http")
async def forward_proto_middleware(request: Request, call_next):
    if request.headers.get("x-forwarded-proto") == "https":
        request.scope["scheme"] = "https"
    response = await call_next(request)
    return response

# Starlette SessionMiddleware (required for Authlib)
app.add_middleware(SessionMiddleware, secret_key=SECRET_KEY)

app.mount("/static", StaticFiles(directory="static"), name="static")
app.mount("/uploads", StaticFiles(directory="uploads"), name="uploads")
templates = Jinja2Templates(directory="templates")

# ======================= UTILS =======================
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def get_user(db: Session, username: str):
    return db.query(User).filter(User.username == username).first()


def get_password_hash(password: str):
    # Bcrypt has a 72-byte limit. Truncate to avoid ValueError.
    return bcrypt.hash(password[:72])


def verify_password(password: str, hashed: str):
    # Bcrypt has a 72-byte limit. Truncate to avoid ValueError.
    return bcrypt.verify(password[:72], hashed)


def create_access_token(data: dict):
    expire = get_ph_time() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    data.update({"exp": expire})
    return jwt.encode(data, SECRET_KEY, algorithm=ALGORITHM)


def calculate_image_similarity(img1_path, img2_path):
    try:
        img1 = Image.open(img1_path).convert("L").resize((100, 100))
        img2 = Image.open(img2_path).convert("L").resize((100, 100))
        a = np.array(img1)
        b = np.array(img2)
        return ssim(a, b, data_range=a.max() - a.min())
    except:
        return 0

import threading

def send_email_notification(recipient_email, subject, message_content):
    """
    Real Email Notification using SMTP (Runs in a background thread to prevent blocking). 
    """
    def send_smtp():
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart
        from email.utils import formatdate, make_msgid
        try:
            msg = MIMEMultipart()
            msg['From'] = MAIL_FROM
            msg['To'] = recipient_email
            msg['Subject'] = subject
            msg['Date'] = formatdate(localtime=True)
            msg['Message-ID'] = make_msgid()
            
            msg.attach(MIMEText(message_content, 'plain'))
            
            server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
            server.starttls()  # Secure the connection
            server.login(SMTP_USER, SMTP_PASS)
            server.send_message(msg)
            server.quit()
            print(f"--- REAL EMAIL SENT TO {recipient_email} ---")
        except Exception as e:
            error_msg = f"--- FAILED TO SEND EMAIL TO {recipient_email}: {e} ---"
            print(error_msg)
            # Log failure to database Notification so admin/user can see it instantly
            try:
                db_session = SessionLocal()
                # Find username associated with this email to notify them
                user_record = db_session.query(User).filter(User.email == recipient_email).first()
                target_user = user_record.username if user_record else recipient_email
                
                notif = Notification(
                    recipient=target_user,
                    message=f"⚠️ SMTP EMAIL FAILURE: Failed to send update email to {recipient_email}. Error: {str(e)}",
                    created_at=get_ph_time()
                )
                db_session.add(notif)
                db_session.commit()
                db_session.close()
            except Exception as db_err:
                print(f"Failed to log email error to database: {db_err}")

    threading.Thread(target=send_smtp, daemon=True).start()

def archive_old_reports(db: Session):
    # Archive items older than 30 days
    expiry_limit = get_ph_time() - timedelta(days=30)
    
    expired_lost = db.query(LostItem).filter(
        LostItem.created_at < expiry_limit, 
        LostItem.is_archived == False
    ).all()
    
    for item in expired_lost:
        item.is_archived = True
        item.status = "Archived (Expired)"
        
    db.commit()


# ======================= STARTUP =======================
@app.on_event("startup")
def startup_events():
    db = SessionLocal()
    try:
        # Create Admin if not exists
        if not get_user(db, "admin"):
            admin = User(
                username="admin",
                first_name="Admin",
                middle_name="",
                last_name="System",
                email="admin@trackbox.com",
                hashed_password=get_password_hash("admin123"),
                role="admin"
            )
            db.add(admin)
            db.commit()
        
        # Create Sample Guard if not exists
        if not get_user(db, "guard"):
            guard = User(
                username="guard",
                first_name="Security",
                middle_name="",
                last_name="Guard",
                email="guard@trackbox.com",
                hashed_password=get_password_hash("guard123"),
                role="guard"
            )
            db.add(guard)
            db.commit()
        
        # Run archiving cleanup
        archive_old_reports(db)
    except Exception as e:
        print(f"Startup task error: {e}")
    finally:
        db.close()

# ======================= AUTH =======================
@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    token = request.cookies.get("access_token")
    if token:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            username = payload.get("sub")
            db = SessionLocal()
            try:
                user = db.query(User).filter(User.username == username).first()
                if user and user.is_active:
                    if payload["role"] == "admin":
                        return RedirectResponse("/admin/dashboard")
                    elif payload["role"] == "guard":
                        return RedirectResponse("/guard/dashboard")
                    else:
                        return RedirectResponse("/student/dashboard")
            finally:
                db.close()
        except:
            pass
    return templates.TemplateResponse("home.html", {"request": request})


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


# ======================= TOGGLE USER STATUS (US-04) =======================
@app.post("/admin/toggle-user-status")
def toggle_user_status(token: str = Form(...), user_id: int = Form(...), db: Session = Depends(get_db)):
    try:
        verify_admin_token(token)
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return JSONResponse({"success": False, "message": "User not found"})
        
        # Admin cannot deactivate themselves (safety measure)
        # But for now we just toggle
        user.is_active = not user.is_active
        db.commit()
        
        status_text = "Activated" if user.is_active else "Deactivated"
        return {"success": True, "message": f"User {user.username} is now {status_text}."}
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

@app.post("/login")
def login(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db)
):
    user = get_user(db, username)
    if not user or not verify_password(password, user.hashed_password):
        return templates.TemplateResponse("login.html", {"request": request, "error": "Invalid username or password"})
    
    if not user.is_active:
        return templates.TemplateResponse("login.html", {"request": request, "error": "Account deactivated. Contact Admin."})

    token = create_access_token({"sub": user.username, "role": user.role})
    
    if user.role == "admin":
        redirect = "/admin/dashboard"
    elif user.role == "guard":
        redirect = "/guard/dashboard"
    else:
        redirect = "/student/dashboard"
    
    # Return JSON for JS handling with SweetAlert
    return JSONResponse(content={"detail": "Login successful!", "redirect": redirect, "token": token})


def render_error_page(title: str, heading: str, message: str, icon: str = "fa-exclamation-triangle", btn_text: str = "Back to Login", btn_url: str = "/login"):
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title} | TrackBox EVSU</title>
    <link href="https://fonts.googleapis.com/css2?family=Poppins:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        :root {{
            --primary: #800000;
            --primary-dark: #5e0000;
            --accent: #FFCC00;
            --bg: #f8fafc;
            --white: #ffffff;
            --text-main: #1e293b;
            --text-muted: #64748b;
        }}
        * {{ box-sizing: border-box; margin: 0; padding: 0; }}
        body {{
            font-family: 'Poppins', sans-serif;
            background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
            color: var(--text-main);
            display: flex;
            align-items: center;
            justify-content: center;
            min-height: 100vh;
            padding: 20px;
        }}
        .container {{
            max-width: 480px;
            width: 100%;
            background: var(--white);
            border-radius: 24px;
            box-shadow: 0 20px 40px rgba(0, 0, 0, 0.08);
            border: 1px solid rgba(0, 0, 0, 0.03);
            text-align: center;
            padding: 40px 30px;
            position: relative;
            overflow: hidden;
            animation: slideUp 0.6s cubic-bezier(0.16, 1, 0.3, 1) forwards;
        }}
        @keyframes slideUp {{
            from {{ opacity: 0; transform: translateY(20px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}
        .brand-header {{
            display: flex;
            align-items: center;
            justify-content: center;
            gap: 10px;
            margin-bottom: 30px;
        }}
        .brand-header img {{
            height: 42px;
        }}
        .brand-header span {{
            font-size: 1.4rem;
            font-weight: 800;
            color: var(--primary);
        }}
        .alert-icon {{
            width: 80px;
            height: 80px;
            border-radius: 50%;
            background: #fee2e2;
            color: #ef4444;
            display: flex;
            align-items: center;
            justify-content: center;
            font-size: 2.2rem;
            margin: 0 auto 25px;
            box-shadow: 0 10px 20px rgba(239, 68, 68, 0.15);
            animation: pulse 2s infinite;
        }}
        @keyframes pulse {{
            0% {{ transform: scale(1); }}
            50% {{ transform: scale(1.05); }}
            100% {{ transform: scale(1); }}
        }}
        h2 {{
            font-size: 1.6rem;
            font-weight: 800;
            color: var(--primary);
            margin-bottom: 12px;
        }}
        p {{
            font-size: 0.95rem;
            color: var(--text-muted);
            line-height: 1.6;
            margin-bottom: 30px;
        }}
        .btn {{
            display: inline-flex;
            align-items: center;
            justify-content: center;
            gap: 10px;
            width: 100%;
            padding: 14px 24px;
            background: var(--primary);
            color: var(--white);
            text-decoration: none;
            font-weight: 600;
            border-radius: 12px;
            transition: all 0.3s ease;
            box-shadow: 0 4px 15px rgba(128, 0, 0, 0.2);
        }}
        .btn:hover {{
            background: var(--primary-dark);
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(128, 0, 0, 0.3);
        }}
        .btn i {{
            font-size: 1rem;
        }}
        .support-info {{
            margin-top: 30px;
            font-size: 0.8rem;
            color: var(--text-muted);
            border-top: 1px solid #f1f5f9;
            padding-top: 20px;
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="brand-header">
            <img src="/static/EVSU_logo.png" alt="EVSU Logo">
            <span>TrackBox</span>
        </div>
        <div class="alert-icon">
            <i class="fas {icon}"></i>
        </div>
        <h2>{heading}</h2>
        <p>{message}</p>
        <a href="{btn_url}" class="btn">
            <i class="fas fa-arrow-left"></i> {btn_text}
        </a>
        <div class="support-info">
            <i class="fas fa-info-circle" style="margin-right: 5px;"></i> Eastern Visayas State University Lost & Found Management System
        </div>
    </div>
</body>
</html>"""


# ======================= OAUTH ROUTES =======================

@app.get("/login/google")
async def login_google(request: Request):
    redirect_uri = request.url_for('auth_google')
    # If not on localhost, force the scheme to https
    if "localhost" not in str(request.base_url) and "127.0.0.1" not in str(request.base_url):
        redirect_uri = str(redirect_uri).replace("http://", "https://")
    return await oauth.google.authorize_redirect(request, str(redirect_uri))

@app.get("/auth/google")
async def auth_google(request: Request, db: Session = Depends(get_db)):
    try:
        token_google = await oauth.google.authorize_access_token(request)
        user_info = token_google.get('userinfo')
        if not user_info:
            return HTMLResponse(render_error_page("Authentication Failed", "Failed to Authenticate", "Failed to fetch user info from Google. Please try again.", icon="fa-user-times"), status_code=400)
        
        email = user_info.get('email')
        
        # Requirement: Must be @evsu.edu.ph
        if not email or not email.lower().endswith("@evsu.edu.ph"):
            return HTMLResponse(render_error_page("Access Denied", "Access Denied", "Only @evsu.edu.ph institutional emails are allowed to register or login.", icon="fa-shield-alt"), status_code=403)

        # Check if user exists
        user = db.query(User).filter(User.email == email).first()
        
        if user and not user.is_active:
            return HTMLResponse(render_error_page("Account Deactivated", "Account Deactivated", "Your institutional or local TrackBox account has been deactivated by the administrator. Please coordinate with campus security post or EVSU administration.", icon="fa-user-slash"), status_code=403)

        if not user:
            # Create new user if not exists
            # We generate a random username if not available, or use the email prefix
            username = email.split('@')[0]
            
            # Check if username exists, if so append random digits
            if get_user(db, username):
                username = f"{username}_{uuid.uuid4().hex[:4]}"
            
            user = User(
                username=username,
                first_name=user_info.get('given_name', 'Student'),
                last_name=user_info.get('family_name', 'EVSU'),
                email=email,
                role="student",
                hashed_password="oauth_user_no_password" # Placeholder
            )
            db.add(user)
            db.commit()
            db.refresh(user)

        # Create JWT token
        token = create_access_token({"sub": user.username, "role": user.role})
        
        # Redirect to dashboard with token in cookie
        if user.role == "admin":
            redirect_url = "/admin/dashboard"
        elif user.role == "guard":
            redirect_url = "/guard/dashboard"
        else:
            redirect_url = "/student/dashboard"
        
        response = RedirectResponse(url=redirect_url)
        response.set_cookie(key="access_token", value=token, httponly=False, max_age=86400)
        return response

    except Exception as e:
        print(f"OAuth Error: {e}")
        return RedirectResponse(url="/login?error=oauth_failed")


@app.get("/logout")
def logout():
    response = RedirectResponse("/", status_code=302)
    response.delete_cookie("access_token")
    return response

# ======================= AUTH REGISTRATION =======================
@app.get("/register", response_class=HTMLResponse)
def register_page(request: Request):
    return templates.TemplateResponse("register.html", {"request": request})

import re

def validate_password_policy(password: str):
    """
    Password Policy:
    At least 1 uppercase, 1 lowercase, 1 number, 1 special character.
    Minimum length: 8 (standard best practice)
    """
    if len(password) < 8:
        return False, "Password must be at least 8 characters long."
    if not re.search(r"[A-Z]", password):
        return False, "Password must contain at least one uppercase letter."
    if not re.search(r"[a-z]", password):
        return False, "Password must contain at least one lowercase letter."
    if not re.search(r"\d", password):
        return False, "Password must contain at least one number."
    if not re.search(r"[!@#$%^&*(),.?\":{}|<>]", password):
        return False, "Password must contain at least one special character."
    return True, ""

@app.post("/register")
def register(
    request: Request,
    username: str = Form(...),
    first_name: str = Form(...),
    middle_name: str = Form(None),
    last_name: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    confirm_password: str = Form(...),
    role: str = Form("student"),
    db: Session = Depends(get_db)
):
    # Password Policy Check
    is_valid, policy_msg = validate_password_policy(password)
    if not is_valid:
        return JSONResponse(status_code=400, content={"detail": policy_msg})

    # Password Confirmation Check
    if password != confirm_password:
        return JSONResponse(status_code=400, content={"detail": "Passwords do not match."})

    # Requirement: Email must be @evsu.edu.ph
    if not email.lower().endswith("@evsu.edu.ph"):
        return JSONResponse(status_code=400, content={"detail": "Registration limited to @evsu.edu.ph emails only."})

    # Logic: Check if username already exists
    if get_user(db, username):
        return JSONResponse(status_code=400, content={"detail": "Username already exists."})
    
    # Logic: Check if email already exists
    existing_email = db.query(User).filter(User.email == email).first()
    if existing_email:
        return JSONResponse(status_code=400, content={"detail": "Email is already registered."})
    
    new_user = User(
        username=username,
        first_name=first_name,
        middle_name=middle_name,
        last_name=last_name,
        email=email,
        hashed_password=get_password_hash(password),
        role=role
    )
    db.add(new_user)
    db.commit()
    return JSONResponse(content={"detail": "Registration successful!"})

# ======================= FORGOT PASSWORD =======================
import uuid

@app.get("/forgot-password", response_class=HTMLResponse)
def forgot_password_page(request: Request):
    return templates.TemplateResponse("forgot_password.html", {"request": request})

@app.post("/forgot-password")
def forgot_password_submit(
    request: Request, 
    email: str = Form(...), 
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.email == email).first()
    if user:
        # Generate a unique token
        token = str(uuid.uuid4())
        user.reset_token = token
        db.commit()
        
        # Real email
        reset_link = f"{request.base_url}reset-password/{token}"
        send_email_notification(
            user.email,
            "TrackBox Password Reset",
            f"Hello {user.first_name},\n\nYou requested a password reset for your TrackBox account. Click the link below to set a new password:\n\n{reset_link}\n\nIf you didn't request this, please ignore this email."
        )
        return JSONResponse(content={"detail": "Password reset link has been sent to your email."})
    else:
        return JSONResponse(status_code=404, content={"detail": "Email address not found."})

@app.get("/reset-password/{token}", response_class=HTMLResponse)
def reset_password_page(request: Request, token: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.reset_token == token).first()
    if not user:
        return RedirectResponse("/login?error=Invalid or expired reset token")
    
    return templates.TemplateResponse("reset_password.html", {"request": request, "token": token})

@app.post("/reset-password/{token}")
def reset_password_submit(
    request: Request,
    token: str,
    password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db)
):
    # Apply Password Policy
    is_valid, policy_msg = validate_password_policy(password)
    if not is_valid:
        return JSONResponse(status_code=400, content={"detail": policy_msg})

    if password != confirm_password:
        return JSONResponse(status_code=400, content={"detail": "Passwords do not match."})
    
    user = db.query(User).filter(User.reset_token == token).first()
    if not user:
        return JSONResponse(status_code=404, content={"detail": "Invalid or expired reset token."})
    
    user.hashed_password = get_password_hash(password)
    user.reset_token = None # Clear the token
    db.commit()
    
    return JSONResponse(content={"detail": "Password updated successfully!"})

@app.get("/student/item-details/{item_id}")
def get_item_details(item_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    
    item = db.query(LostItem).filter(LostItem.id == item_id).first()
    if not item:
        item = db.query(FoundItem).filter(FoundItem.id == item_id).first()
    
    if not item:
        return JSONResponse({"success": False, "message": "Item not found"}, status_code=404)
    
    is_lost_table = isinstance(item, LostItem)
    
    # If there's a match, get those details too for comparison
    match_details = None
    if is_lost_table and item.matched_with:
        match = db.query(LostItem).filter(LostItem.id == item.matched_with).first()
        if match:
            match_details = {
                "id": match.id,
                "item_name": match.item_name,
                "category": match.category,
                "description": match.description,
                "reporter": match.reporter,
                "type": (match.type or "Lost").capitalize(),
                "location": match.last_seen,
                "image": match.image,
                "created_at": match.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                "status": match.status
            }

    return {
        "success": True,
        "item": {
            "id": item.id,
            "item_name": item.item_name,
            "category": item.category,
            "description": item.description,
            "reporter": item.reporter,
            "type": (item.type or "Lost").capitalize() if is_lost_table else "Found",
            "location": item.last_seen if is_lost_table else item.found_in,
            "image": item.image,
            "created_at": item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            "status": item.status,
            "is_owner_verified": getattr(item, 'is_owner_verified', False) if is_lost_table else False,
            "verified_at": item.verified_at.strftime('%Y-%m-%d %H:%M:%S') if is_lost_table and item.verified_at else None
        },
        "match": match_details
    }

@app.post("/student/verify-match/{item_id}")
async def verify_match(item_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    
    # The user verifying must be the one who reported it as LOST
    item = db.query(LostItem).filter(LostItem.id == item_id, LostItem.reporter == user.username).first()
    if not item:
        return JSONResponse({"success": False, "message": "Item report not found or unauthorized"}, status_code=404)
    
    if item.type.upper() != "LOST":
        return JSONResponse({"success": False, "message": "Only lost item reports can be verified by the owner"}, status_code=400)
    
    if not item.matched_with:
        return JSONResponse({"success": False, "message": "No match found to verify"}, status_code=400)
    
    # Mark as verified
    item.is_owner_verified = True
    item.verified_at = get_ph_time()
    item.status = "Verified by Owner"
    
    # Also update the matched item (the one reported as FOUND)
    matched_item = db.query(LostItem).filter(LostItem.id == item.matched_with).first()
    if matched_item:
        matched_item.status = "Verified by Owner"
        matched_item.is_owner_verified = True # Mark both as verified for consistency
        matched_item.verified_at = item.verified_at
        
        # Notify the finder
        db.add(Notification(
            recipient=matched_item.reporter,
            message=f"Good news! The owner of '{item.item_name}' has verified that the item you found is theirs. You can now coordinate the return!"
        ))
    
    db.commit()
    return {"success": True, "message": "Match verified successfully!", "verified_at": item.verified_at.strftime('%Y-%m-%d %H:%M:%S')}


# ======================= DASHBOARDS =======================
@app.get("/admin/dashboard", response_class=HTMLResponse)
def admin_dashboard(request: Request, db: Session = Depends(get_db)):
    auth_token = request.cookies.get("access_token")
    if not auth_token:
        return RedirectResponse("/login")
    try:
        verify_admin_token(auth_token)
    except HTTPException:
        return RedirectResponse("/login")
    
    # Accurate Counts
    # Lost reports that are still active
    lost_count = db.query(LostItem).filter(LostItem.type == "LOST", LostItem.is_archived == False).count()
    
    # Found reports (from students) + Found items (official) that are not claimed
    found_count = db.query(LostItem).filter(LostItem.type == "FOUND", LostItem.is_archived == False).count() + \
                  db.query(FoundItem).filter(FoundItem.is_claimed == False).count()
    
    # Total Returned (Archived LostItems + Claimed FoundItems)
    returned_count = db.query(LostItem).filter(LostItem.is_archived == True).count() + \
                    db.query(FoundItem).filter(FoundItem.is_claimed == True).count()
    
    # Fetch admin notifications
    notifications = db.query(Notification).filter(Notification.recipient == "admin").order_by(Notification.created_at.desc()).limit(10).all()
    unread_count = db.query(Notification).filter(Notification.recipient == "admin", Notification.is_read == False).count()

    return templates.TemplateResponse("dashboard_admin.html", {
        "request": request, 
        "token": auth_token,
        "lost_count": lost_count,
        "found_count": found_count,
        "returned_count": returned_count,
        "notifications": notifications,
        "unread_count": unread_count
    })

@app.get("/student/dashboard", response_class=HTMLResponse)
@app.get("/dashboard_student", response_class=HTMLResponse)
def dashboard_student(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login")
    
    # Fetch notifications
    notifications = db.query(Notification).filter(Notification.recipient == user.username).order_by(Notification.created_at.desc()).limit(10).all()
    unread_count = db.query(Notification).filter(Notification.recipient == user.username, Notification.is_read == False).count()
    
    # Fetch recent lost items (only verified/approved items visible to the public)
    lost_items = db.query(LostItem).filter(
        LostItem.reporter != user.username, 
        LostItem.is_archived == False,
        LostItem.status.in_(["Approved", "Verified", "Active", "IN CUSTODY", "Potential Match Found"])
    ).order_by(LostItem.created_at.desc()).limit(10).all()

    return templates.TemplateResponse("dashboard_student.html", {
        "request": request,
        "user": user,
        "lost_items": lost_items,
        "notifications": notifications,
        "unread_count": unread_count
    })

@app.get("/report_lost", response_class=HTMLResponse)
def report_lost_page_alias(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login")
    
    token = request.cookies.get("access_token")
    return templates.TemplateResponse("report_lost.html", {
        "request": request,
        "user": user,
        "token": token
    })

@app.post("/report_lost")
async def report_lost_post_alias(
    request: Request,
    token: str = Form(...),
    type: str = Form(...),
    category: str = Form(...),
    item_name: str = Form(...),
    description: str = Form(...),
    last_seen: str = Form(...),
    building: str = Form(...),
    security_question: Optional[str] = Form(None),
    security_answer: Optional[str] = Form(None),
    image: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    return await report_lost(
        request, token, type, category, item_name, description, 
        last_seen, building, security_question, security_answer, image, db
    )

@app.post("/user_account/change-password")
def change_password(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db)
):
    token = request.cookies.get("access_token")
    if not token:
        return RedirectResponse("/login")

    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user = get_user(db, payload["sub"])
    except:
        return RedirectResponse("/login")

    # Check current password
    if not verify_password(current_password, user.hashed_password):
        return templates.TemplateResponse(
            "user_account.html",
            {"request": request, "user": user, "error": "Current password is incorrect"}
        )

    # Check new password confirmation
    if new_password != confirm_password:
        return templates.TemplateResponse(
            "user_account.html",
            {"request": request, "user": user, "error": "New passwords do not match"}
        )

    # Update password
    user.hashed_password = get_password_hash(new_password)
    db.commit()

    return templates.TemplateResponse(
        "user_account.html",
        {"request": request, "user": user, "success": "Password changed successfully!"}
    )

# ======================= ADMIN DASHBOARD =======================
USERS = [{"id": 1, "username": "{{ user.username }}", "email": "student01@example.com", "role": "student"}]

def get_current_user(request: Request, db: Session):
    token = request.cookies.get("access_token")
    if not token:
        return None
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        user = db.query(User).filter(User.username == username).first()
        if user and not user.is_active:
            return None
        return user
    except:
        return None

@app.get("/user_account", response_class=HTMLResponse)
def user_account(request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login")

    # Fetch user posts (lost + found items reported by user) — including archived so we can show returned items
    user_lost = db.query(LostItem).filter(LostItem.reporter == user.username).order_by(LostItem.created_at.desc()).all()
    user_found = db.query(FoundItem).filter(FoundItem.reporter == user.username).order_by(FoundItem.created_at.desc()).all()

    posts = []
    completed_posts = []

    # Process Lost Items
    for item in user_lost:
        if item.status == "Rejected":
            continue

        is_returned = any(s in (item.status or "").upper() for s in ["RETURNED", "CLAIMED", "COMPLETED", "SUCCESSFUL"]) or item.is_archived
        
        matched_user = None
        match = None
        chat_id = item.id
        if item.matched_with:
            chat_id = min(item.id, item.matched_with)
            match = db.query(LostItem).filter(LostItem.id == item.matched_with).first()
            if match:
                matched_user = match.reporter

        post_data = {
            "id": item.id,
            "chat_id": chat_id,
            "title": item.item_name,
            "type": item.type or "Lost",
            "description": item.description,
            "status": item.status,
            "matched_with": item.matched_with,
            "matched_user": matched_user,
            "match_details": {
                "item_name": match.item_name if match else None,
                "description": match.description if match else None,
                "image": match.image if match else None,
                "location": match.last_seen if match else None
            } if match else None,
            "is_owner_verified": item.is_owner_verified,
            "verified_at": item.verified_at.strftime('%Y-%m-%d %H:%M:%S') if item.verified_at else None,
            "is_claimed": is_returned,
            "handed_over_by": item.handed_over_by,
            "handover_method": item.handover_method,
            "claimed_at": item.claimed_at.strftime('%Y-%m-%d %H:%M:%S') if item.claimed_at else None,
            "created_at": item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else None
        }

        # Normalize status
        if post_data["status"] == "Completed" or is_returned:
            post_data["status"] = "RETURNED (PEER TO PEER)"

        if is_returned:
            completed_posts.append(post_data)
        else:
            posts.append(post_data)

    # Process Found Items
    for item in user_found:
        if item.status == "Rejected":
            continue

        is_returned = item.is_claimed or any(s in (item.status or "").upper() for s in ["RETURNED", "CLAIMED", "COMPLETED", "SUCCESSFUL"]) or item.is_archived

        post_data = {
            "id": item.id,
            "title": item.item_name,
            "type": "Found",
            "description": item.description,
            "status": item.status,
            "matched_with": None,
            "is_claimed": is_returned,
            "handed_over_by": item.handed_over_by,
            "handed_over_to": item.handed_over_to,
            "handover_method": item.handover_method,
            "claimed_at": item.claimed_at.strftime('%Y-%m-%d %H:%M:%S') if item.claimed_at else None,
            "created_at": item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else None
        }

        # Normalize status
        if post_data["status"] == "Completed" or is_returned:
            post_data["status"] = "RETURNED (PEER TO PEER)"

        if is_returned:
            completed_posts.append(post_data)
        else:
            posts.append(post_data)

    # Sort both lists: newest first
    posts.sort(key=lambda x: x["created_at"] or "", reverse=True)
    completed_posts.sort(key=lambda x: x["claimed_at"] or x["created_at"] or "", reverse=True)

    # Fetch notifications
    notifications = db.query(Notification).filter(Notification.recipient == user.username).order_by(Notification.created_at.desc()).limit(10).all()
    unread_count = db.query(Notification).filter(Notification.recipient == user.username, Notification.is_read == False).count()

    return templates.TemplateResponse("user_account.html", {
        "request": request,
        "user": user,
        "token": request.cookies.get("access_token"),
        "posts": posts,
        "completed_posts": completed_posts,
        "notifications": notifications,
        "unread_count": unread_count
    })

@app.post("/user/upload_profile_image")
async def upload_profile_image(
    request: Request,
    profile_image: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login")
    
    # Handle file upload
    if profile_image:
        timestamp = int(get_ph_time().timestamp())
        filename = f"profile_{user.id}_{timestamp}_{profile_image.filename}"
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        with open(file_path, "wb") as f:
            shutil.copyfileobj(profile_image.file, f)
        
        user.profile_image = filename
        db.commit()
    
    return RedirectResponse("/user_account", status_code=303)

@app.get("/edit_report/{item_id}", response_class=HTMLResponse)
def edit_report_page(item_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login")
    
    item = db.query(LostItem).filter(LostItem.id == item_id, LostItem.reporter == user.username).first()
    if not item:
        item = db.query(FoundItem).filter(FoundItem.id == item_id, FoundItem.reporter == user.username).first()
    
    if not item:
        raise HTTPException(status_code=404, detail="Item not found or unauthorized")
        
    return templates.TemplateResponse("edit_post.html", {"request": request, "post": item, "user": user})

@app.post("/edit_post/{item_id}")
async def edit_post(
    item_id: int,
    request: Request,
    title: str = Form(...),
    description: str = Form(...),
    type: str = Form(...),
    image: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login")
    
    # Try item in LostItem
    item = db.query(LostItem).filter(LostItem.id == item_id, LostItem.reporter == user.username).first()
    is_lost = True
    if not item:
        item = db.query(FoundItem).filter(FoundItem.id == item_id, FoundItem.reporter == user.username).first()
        is_lost = False
    
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")

    item.item_name = title
    item.description = description
    
    if image and image.filename:
        timestamp = int(datetime.utcnow().timestamp())
        filename = f"{timestamp}_{image.filename}"
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        with open(file_path, "wb") as f:
            shutil.copyfileobj(image.file, f)
        item.image = filename

    db.commit()
    return RedirectResponse("/user_account", status_code=303)

@app.post("/delete_report/{item_id}")
def delete_report(item_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login")
        
    item = db.query(LostItem).filter(LostItem.id == item_id, LostItem.reporter == user.username).first()
    if not item:
        item = db.query(FoundItem).filter(FoundItem.id == item_id, FoundItem.reporter == user.username).first()
        
    if item:
        db.delete(item)
        db.commit()
        
    return RedirectResponse("/user_account", status_code=303)

@app.get("/search_lost", response_class=HTMLResponse)
def search_lost_alias(request: Request, db: Session = Depends(get_db)):
    return student_search(request, None, db)

@app.post("/search_lost")
async def search_lost_post_alias(
    request: Request,
    token: str = Form(...),
    query: str = Form(""),
    image: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    return await search_items_post(request, token, query, image, db)

# ======================= ADMIN LOST ITEMS =======================
def verify_admin_token(token: str):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        role = payload.get("role")
        if role not in ["admin", "guard"]:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized")
        
        # Verify user is active in database
        username = payload.get("sub")
        db = SessionLocal()
        try:
            user = db.query(User).filter(User.username == username).first()
            if not user or not user.is_active:
                raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not authorized or deactivated")
        finally:
            db.close()
            
        return username
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid token")

# --- Endpoints ---

@app.get("/admin/lost-items", response_class=HTMLResponse)
def admin_lost_items(request: Request, token: Optional[str] = None, db: Session = Depends(get_db)):
    # STRICT authetication via cookie
    auth_token = request.cookies.get("access_token")
    if not auth_token:
         return RedirectResponse("/login")

    try:
        verify_admin_token(auth_token)
    except HTTPException:
        return RedirectResponse("/login")

    notifications = db.query(Notification).filter(Notification.recipient == "admin").order_by(Notification.created_at.desc()).limit(10).all()
    unread_count = db.query(Notification).filter(Notification.recipient == "admin", Notification.is_read == False).count()

    response = templates.TemplateResponse("admin_lost_items.html", {
        "request": request,
        "token": auth_token,
        "notifications": notifications,
        "unread_count": unread_count
    })
    # Prevent browser from caching this page
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.get("/admin/found-items", response_class=HTMLResponse)
def admin_found_items(request: Request, token: Optional[str] = None, db: Session = Depends(get_db)):
    # STRICT authetication via cookie
    auth_token = request.cookies.get("access_token")
    if not auth_token:
         return RedirectResponse("/login")

    try:
        verify_admin_token(auth_token)
    except HTTPException:
        return RedirectResponse("/login")

    found_items = db.query(FoundItem).order_by(FoundItem.created_at.desc()).all()
    
    notifications = db.query(Notification).filter(Notification.recipient == "admin").order_by(Notification.created_at.desc()).limit(10).all()
    unread_count = db.query(Notification).filter(Notification.recipient == "admin", Notification.is_read == False).count()
    
    response = templates.TemplateResponse("admin_found_items.html", {
        "request": request, 
        "token": auth_token,
        "found_items": found_items,
        "notifications": notifications,
        "unread_count": unread_count
    })
    # Prevent browser from caching this page
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.get("/items/all")
def get_all_items(token: str, db: Session = Depends(get_db)):
    try:
        username = verify_admin_token(token)
    except HTTPException as e:
        return JSONResponse(status_code=e.status_code, content={"detail": e.detail})

    # Fetch user to check role and assigned building
    user = db.query(User).filter(User.username == username).first()
    building_filter = user.assigned_building if (user and user.role == "guard") else None

    if user and user.role == "guard":
        # Guards should only see active, unclaimed items that are strictly in custody
        found_q = db.query(FoundItem).filter(
            FoundItem.is_archived == False,
            FoundItem.is_claimed == False,
            FoundItem.status == "IN CUSTODY"
        )
        if building_filter:
            found_q = found_q.filter(FoundItem.building == building_filter)
        found_items = found_q.order_by(FoundItem.created_at.desc()).all()
        lost_items = []
        found_in_lost_table = []
    else:
        # Get LOST items from lost_items table
        lost_q = db.query(LostItem).filter(
            LostItem.is_archived == False,
            LostItem.type == "LOST"
        )
        if building_filter:
            lost_q = lost_q.filter(LostItem.building == building_filter)
        lost_items = lost_q.order_by(LostItem.created_at.desc()).all()
        
        # Get FOUND items from found_items table
        found_q = db.query(FoundItem).filter(FoundItem.is_archived == False)
        if building_filter:
            found_q = found_q.filter(FoundItem.building == building_filter)
        found_items = found_q.order_by(FoundItem.created_at.desc()).all()
        
        # Get FOUND items from lost_items table (unified reporting)
        # Include peer-to-peer items even if archived for admin visibility
        found_lost_q = db.query(LostItem).filter(
            LostItem.type == "FOUND"
        )
        if building_filter:
            found_lost_q = found_lost_q.filter(LostItem.building == building_filter)
        found_in_lost_table = found_lost_q.order_by(LostItem.created_at.desc()).all()
        
        # Filter: show non-archived OR peer-to-peer claimed items
        found_in_lost_table = [
            item for item in found_in_lost_table 
            if not item.is_archived or "Peer-to-Peer" in (item.status or "")
        ]

    all_items = []
    
    # Add LOST items
    for item in lost_items:
        all_items.append({
            "id": item.id,
            "item_name": item.item_name,
            "category": item.category,
            "image": item.image,
            "location": item.last_seen,
            "reported_at": item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            "type": "lost",
            "status": item.status,
            "is_verified": False,
            "is_claimed": False,
            "reporter": item.reporter
        })
    
    # Add FOUND items from found_items table
    for item in found_items:
        all_items.append({
            "id": item.id,
            "item_name": item.item_name,
            "category": item.category,
            "image": item.image,
            "location": item.found_in,
            "reported_at": item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            "type": "found",
            "status": item.status,
            "is_verified": item.is_verified,
            "is_claimed": item.is_claimed,
            "reporter": item.reporter
        })
    
    # Add FOUND items from lost_items table
    for item in found_in_lost_table:
        # Check if this item has been claimed via peer-to-peer
        is_p2p_claimed = "Peer-to-Peer" in (item.status or "")
        
        all_items.append({
            "id": item.id,
            "item_name": item.item_name,
            "category": item.category,
            "image": item.image,
            "location": item.last_seen,
            "reported_at": item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            "type": "found",
            "status": item.status,
            "is_verified": False,  # Student custody
            "is_claimed": is_p2p_claimed,
            "reporter": item.reporter
        })

    all_items.sort(key=lambda x: x["reported_at"], reverse=True)
    return JSONResponse(content=all_items)

@app.delete("/admin/delete/{item_id}")
def delete_item(item_id: int, token: str, table: Optional[str] = None, db: Session = Depends(get_db)):
    try:
        verify_admin_token(token)
    except HTTPException as e:
        return JSONResponse(status_code=e.status_code, content={"detail": e.detail})

    item = None
    # Look up in the requested table first, with fallback to the other table
    if table == "found":
        item = db.query(FoundItem).filter(FoundItem.id == item_id).first()
        if not item:
            item = db.query(LostItem).filter(LostItem.id == item_id).first()
    elif table == "lost":
        item = db.query(LostItem).filter(LostItem.id == item_id).first()
        if not item:
            item = db.query(FoundItem).filter(FoundItem.id == item_id).first()
    else:
        # Fallback search both
        item = db.query(LostItem).filter(LostItem.id == item_id).first()
        if not item:
            item = db.query(FoundItem).filter(FoundItem.id == item_id).first()

    if not item:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")

    # Soft Delete (Archiving) instead of hard delete
    item.is_archived = True
    item.status = "Rejected"
    db.commit()

    # Create in-app Rejection Notification for the student dashboard
    try:
        if item.reporter:
            notif = Notification(
                recipient=item.reporter,
                message=f"❌ Your report for the item '{item.item_name}' has been reviewed by the Admin and was REJECTED.",
                created_at=get_ph_time()
            )
            db.add(notif)
            db.commit()
    except Exception as notif_err:
        print(f"Failed to create rejection notification: {notif_err}")

    return {"success": True, "message": "Item archived successfully and reporter notified"}
    
@app.get("/admin/qr-scanner", response_class=HTMLResponse)
def qr_scanner(request: Request, token: Optional[str] = None):
    # STRICT authetication via cookie
    auth_token = request.cookies.get("access_token")

    if not auth_token:
         return RedirectResponse("/login")

    try:
        payload = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
        role = payload.get("role")
        username = payload.get("sub")
        if role not in ["admin", "guard"]:
            return RedirectResponse("/login")
        
        # Get user for template
        from sqlalchemy.orm import Session
        db: Session = next(get_db())
        user = db.query(User).filter(User.username == username).first()

        template_name = "admin_qr_scanner.html" if role == "admin" else "guard_qr_scanner.html"
        response = templates.TemplateResponse(template_name, {"request": request, "token": auth_token, "user": user})
    except Exception:
        return RedirectResponse("/login")
        
    # Prevent browser from caching this page
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.get("/admin/users")
def get_users(token: str, db: Session = Depends(get_db)):
    try:
        verify_admin_token(token)
    except:
        return JSONResponse(status_code=401, content={"detail": "Unauthorized"})
    
    users = db.query(User).all()
    return [{"id": u.id, "username": u.username, "email": u.email, "role": u.role, "assigned_building": u.assigned_building, "is_active": u.is_active} for u in users]

@app.post("/admin/create-user")
def admin_create_user(
    token: str = Form(...),
    username: str = Form(...),
    first_name: str = Form(...),
    last_name: str = Form(...),
    email: str = Form(...),
    password: str = Form(...),
    role: str = Form(...),
    db: Session = Depends(get_db)
):
    try:
        verify_admin_token(token)
        
        if get_user(db, username):
            return JSONResponse({"success": False, "message": "Username already exists"})
        
        if db.query(User).filter(User.email == email).first():
            return JSONResponse({"success": False, "message": "Email already exists"})
        
        new_user = User(
            username=username,
            first_name=first_name,
            last_name=last_name,
            email=email,
            hashed_password=get_password_hash(password),
            role=role
        )
        db.add(new_user)
        db.commit()
        
        return JSONResponse({"success": True, "message": f"Account for {username} created successfully as {role}."})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.post("/admin/change-role")
def change_role(
    token: str = Form(...),
    user_id: int = Form(...),
    new_role: str = Form(...),
    assigned_building: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    try:
        verify_admin_token(token)
        user = db.query(User).filter(User.id == user_id).first()
        if not user:
            return JSONResponse({"success": False, "message": "User not found"})
        
        old_role = user.role
        user.role = new_role
        user.assigned_building = assigned_building
        db.commit()
        
        # Log action for audit (US-07)
        db.add(Notification(
            recipient="admin",
            message=f"RBAC Update: User {user.username} role changed from {old_role} to {new_role}. Assigned Building: {assigned_building or 'None'}."
        ))
        db.commit()
        
        return JSONResponse({"success": True, "message": f"User {user.username} updated to {new_role} ({assigned_building or 'No specific building'})."})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

# Add this endpoint to verify QR codes
@app.post("/admin/verify-qr")
async def verify_qr(request: Request, token: str = Form(None), qr_data: str = Form(...), db: Session = Depends(get_db)):
    try:
        # Accept token from form body OR fallback to session cookie
        auth_token = token or request.cookies.get("access_token")
        if not auth_token:
            raise HTTPException(status_code=401, detail="No token")
        payload = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
        role = payload.get("role")
        if role not in ["admin", "guard"]:
            raise HTTPException(status_code=403, detail="Not authorized")
    except Exception:
        raise HTTPException(status_code=401, detail="Invalid token")

    # Parse QR data (format: "CLAIM ITEM {item_id} BY {username}")
    if not qr_data.startswith("CLAIM ITEM"):
        return JSONResponse({"success": False, "message": "Invalid QR code format"})

    try:
        # Extract item ID
        parts = qr_data.split()
        item_id = int(parts[2])

        # Try to find in LostItem first
        item = db.query(LostItem).filter(LostItem.id == item_id).first()

        # If not found in LostItem, try FoundItem
        if not item:
            item = db.query(FoundItem).filter(FoundItem.id == item_id).first()

        if not item:
            return JSONResponse({"success": False, "message": "Item not found"})

        # Return item details
        is_lost_table = isinstance(item, LostItem)
        actual_type = (item.type or "lost").lower() if is_lost_table else "found"
        # We consider it verified if it's in the FoundItem table
        is_verified = not is_lost_table

        item_details = {
            "id": item.id,
            "item_name": item.item_name,
            "category": item.category,
            "description": item.description,
            "reporter": item.reporter,
            "type": actual_type,
            "location": item.last_seen if is_lost_table else item.found_in,
            "image": item.image,
            "reported_at": item.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            "is_claimed": getattr(item, 'is_claimed', False),
            "is_verified": is_verified
        }

        return JSONResponse({"success": True, "item": item_details})

    except Exception as e:
        return JSONResponse({"success": False, "message": f"Error verifying QR: {str(e)}"})

# ======================= ADMIN VERIFICATION =======================

@app.post("/admin/approve-item/{item_type}/{item_id}")
async def approve_item(item_type: str, item_id: int, token: str = Form(...), db: Session = Depends(get_db)):
    try:
        verify_admin_token(token)
    except:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    
    item = None
    if item_type == "lost":
        item = db.query(LostItem).filter(LostItem.id == item_id).first()
        if not item:
            item = db.query(FoundItem).filter(FoundItem.id == item_id).first()
    else:
        item = db.query(FoundItem).filter(FoundItem.id == item_id).first()
        if not item:
            item = db.query(LostItem).filter(LostItem.id == item_id).first()
        
    if not item:
        return JSONResponse({"success": False, "message": "Item not found"})
        
    item.status = "Approved"
    
    # Notify reporter
    notif = Notification(
        recipient=item.reporter,
        message=f"Your {item_type} report for '{item.item_name}' has been approved and is now visible to others.",
        created_at=get_ph_time()
    )
    db.add(notif)
    db.commit()
    
    # RUN MATCHING NOW
    await run_matching(db, item_id, item_type)
    
    return JSONResponse({"success": True, "message": "Report approved and matching checked"})

# ======================= GUARD DASHBOARD =======================

@app.get("/guard/qr-scanner", response_class=HTMLResponse)
def guard_qr_scanner_page(request: Request, db: Session = Depends(get_db)):
    """Dedicated Verification Scanner for guards only (cookie-based auth)."""
    auth_token = request.cookies.get("access_token")
    if not auth_token:
        return RedirectResponse("/login")
    try:
        payload = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
        role = payload.get("role")
        username = payload.get("sub")
        if role != "guard":
            return RedirectResponse("/login")
        user = db.query(User).filter(User.username == username).first()
    except Exception:
        return RedirectResponse("/login")

    response = templates.TemplateResponse("guard_qr_scanner.html", {
        "request": request,
        "token": auth_token,
        "user": user
    })
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response

@app.get("/guard/dashboard", response_class=HTMLResponse)
def guard_dashboard(request: Request, db: Session = Depends(get_db)):
    auth_token = request.cookies.get("access_token")
    if not auth_token:
        return RedirectResponse("/login")
    try:
        username = verify_admin_token(auth_token) # This returns username if valid
        user = db.query(User).filter(User.username == username).first()
    except HTTPException:
        return RedirectResponse("/login")
    
    # Stats for guard's building if assigned
    building_filter = user.assigned_building
    
    found_q = db.query(FoundItem).filter(FoundItem.is_claimed == False)
    if building_filter:
        found_q = found_q.filter(FoundItem.building == building_filter)
        
    found_count = found_q.count()
    notifications = db.query(Notification).filter(Notification.recipient == "admin").order_by(Notification.created_at.desc()).limit(10).all()

    return templates.TemplateResponse("dashboard_guard.html", {
        "request": request, 
        "user": user,
        "token": auth_token,
        "found_count": found_count,
        "notifications": notifications,
        "assigned_building": building_filter,
        "all_buildings": db.query(Building).all()
    })


@app.get("/guard/inventory", response_class=HTMLResponse)
def guard_inventory_page(request: Request, db: Session = Depends(get_db)):
    auth_token = request.cookies.get("access_token")
    if not auth_token:
        return RedirectResponse("/login")
    try:
        username = verify_admin_token(auth_token)
        user = db.query(User).filter(User.username == username).first()
    except HTTPException:
        return RedirectResponse("/login")
    
    if user.role != "guard":
        return RedirectResponse("/login")
        
    building_filter = user.assigned_building
    found_q = db.query(FoundItem).filter(FoundItem.is_claimed == False)
    if building_filter:
        found_q = found_q.filter(FoundItem.building == building_filter)
    found_count = found_q.count()

    response = templates.TemplateResponse("guard_inventory.html", {
        "request": request,
        "token": auth_token,
        "user": user,
        "found_count": found_count,
        "assigned_building": building_filter
    })
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.post("/guard/quick-drop")
async def quick_drop(
    request: Request,
    token: str = Form(...),
    finder_id: str = Form(...),
    item_name: str = Form(...),
    category: str = Form(...),
    description: str = Form(...),
    found_in: str = Form(...),
    building: str = Form(...),
    image: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    try:
        verify_admin_token(token)
        
        # Verify finder exists (Institutional account)
        finder_id_clean = finder_id.strip()
        
        # 1. Exact match on username or email (case-insensitive)
        finder = db.query(User).filter(
            (func.lower(User.username) == finder_id_clean.lower()) | 
            (func.lower(User.email) == finder_id_clean.lower())
        ).first()
        
        # 2. Match by First Name + Last Name (case-insensitive)
        if not finder:
            finder = db.query(User).filter(
                func.lower(func.concat(User.first_name, " ", User.last_name)) == finder_id_clean.lower()
            ).first()
            
        # 3. Match by First Name + Middle Name + Last Name (case-insensitive)
        if not finder:
            finder = db.query(User).filter(
                func.lower(func.concat(User.first_name, " ", User.middle_name, " ", User.last_name)) == finder_id_clean.lower()
            ).first()
            
        # 4. Match by first_name only (if unique)
        if not finder:
            matches = db.query(User).filter(func.lower(User.first_name) == finder_id_clean.lower()).all()
            if len(matches) == 1:
                finder = matches[0]
                
        # 5. Match by last_name only (if unique)
        if not finder:
            matches = db.query(User).filter(func.lower(User.last_name) == finder_id_clean.lower()).all()
            if len(matches) == 1:
                finder = matches[0]
                
        # 6. Match by username containing search string (if unique)
        if not finder:
            matches = db.query(User).filter(User.username.ilike(f"%{finder_id_clean}%")).all()
            if len(matches) == 1:
                finder = matches[0]

        if not finder:
             return JSONResponse({"success": False, "message": "Finder account not found. Must be a valid EVSU user."})

        # Handle image
        filename = None
        if image and image.filename:
            timestamp = int(get_ph_time().timestamp())
            filename = f"found_{timestamp}_{image.filename}"
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            with open(file_path, "wb") as f:
                shutil.copyfileobj(image.file, f)

        # Create Found Item with status IN CUSTODY (Philippine Time)
        new_item = FoundItem(
            reporter=finder.username,
            category=category,
            item_name=item_name,
            description=description,
            found_in=found_in,
            building=building,
            image=filename,
            status="IN CUSTODY",
            created_at=get_ph_time()
        )
        db.add(new_item)
        db.commit()

        # Notify Finder
        db.add(Notification(
            recipient=finder.username,
            message=f"Thank you! Your found item '{item_name}' has been recorded and is now IN CUSTODY of the security office."
        ))
        db.commit()

        # Check for matching lost items reported by students to notify them
        matching_lost_items = db.query(LostItem).filter(
            LostItem.category == category,
            LostItem.type.ilike("LOST"),
            LostItem.is_archived == False,
            LostItem.status.in_(["Approved", "Verified", "Active", "Pending", "Potential Match Found"])
        ).all()

        for lost_item in matching_lost_items:
            # Case-insensitive substring match
            if item_name.lower() in lost_item.item_name.lower() or lost_item.item_name.lower() in item_name.lower():
                # Update status of student's lost report to let them know it's in guard's custody
                lost_item.status = "IN CUSTODY"
                
                # Add system notification for the student
                db.add(Notification(
                    recipient=lost_item.reporter,
                    message=f"Great news! A matching item for your lost report '{lost_item.item_name}' was turned in and is now IN CUSTODY at the '{building}' Guard Station. You can claim it by showing your Claim QR code to the guard!"
                ))
                
                # Send WebSocket notification (Toast Alert)
                try:
                    await manager.send_personal_message({
                        "type": "notification",
                        "message": f"Great news! Your lost '{lost_item.item_name}' is now IN CUSTODY at the '{building}' Guard Station! 🛡️"
                    }, lost_item.reporter)
                except:
                    pass

                # Email Notification
                student_user = db.query(User).filter(User.username.ilike(lost_item.reporter)).first()
                if not student_user:
                    # Fallback: search by email prefix
                    student_user = db.query(User).filter(User.email.ilike(f"{lost_item.reporter}@%")).first()
                
                if student_user and student_user.email:
                    try:
                        send_email_notification(
                            recipient_email=student_user.email,
                            subject="Your Lost Item is in Custody!",
                            message_content=f"Hello {student_user.first_name},\n\nGood news! A matching item for your reported lost item '{lost_item.item_name}' has been recorded and is now in custody at the '{building}' Guard Station.\n\nPlease visit the guard house to claim it using your Claim QR code.\n\nBest regards,\nTrackBox Team"
                        )
                    except:
                        pass
        db.commit()

        return JSONResponse({"success": True, "message": f"Item '{item_name}' successfully recorded as IN CUSTODY."})
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)
    
# ======================= REPORT LOST =======================
# ================= GET ROUTE =================
@app.get("/student/report-lost", response_class=HTMLResponse)
def report_lost_page(request: Request, token: str):
    try:
        data = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if data["role"] != "student":
            raise Exception()
    except:
        return RedirectResponse("/login")

    response = templates.TemplateResponse("report_lost.html", {"request": request, "token": token})
    
    # Prevent browser caching
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    
    return response

# ================= POST ROUTE =================
@app.post("/student/report-lost", response_class=HTMLResponse)
async def report_lost(
    request: Request,
    token: str = Form(...),
    type: str = Form(...),          # NEW: type (LOST or FOUND)
    category: str = Form(...),
    item_name: str = Form(...),
    description: str = Form(...),
    last_seen: str = Form(...),
    building: str = Form(...),
    security_question: Optional[str] = Form(None),
    security_answer: Optional[str] = Form(None),
    image: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    # Verify token
    try:
        data = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        reporter = data["sub"]
    except:
        return RedirectResponse("/login")
    
    # Handle file upload
    filename = None
    if image and image.filename:
        # Ensure upload folder exists
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        timestamp = int(get_ph_time().timestamp())
        filename = f"{timestamp}_{image.filename}"
        file_path = os.path.join(UPLOAD_FOLDER, filename)
        with open(file_path, "wb") as f:
            shutil.copyfileobj(image.file, f)
    
    # Create LostItem instance
    item = LostItem(
        reporter=reporter,
        type=type,
        category=category,
        item_name=item_name,
        description=description,
        last_seen=last_seen,
        building=building,
        image=filename
    )

    # Handle security question for FOUND items
    if type.upper() == "FOUND":
        item.security_question = security_question
        item.security_answer = security_answer
    
    # Save to DB
    db.add(item)
    db.commit()
    db.refresh(item)
    return templates.TemplateResponse("report_lost.html", {
        "request": request,
        "token": token,
        "success": "Report submitted! Awaiting admin verification before it goes public.",
        "user": get_user(db, reporter)
    })

async def run_matching(db: Session, item_id: int, item_type: str):
    """Utility to run matching process after an item is approved."""
    item = None
    if item_type == "lost":
        item = db.query(LostItem).filter(LostItem.id == item_id).first()
        if not item:
            item = db.query(FoundItem).filter(FoundItem.id == item_id).first()
    else:
        item = db.query(FoundItem).filter(FoundItem.id == item_id).first()
        if not item:
            item = db.query(LostItem).filter(LostItem.id == item_id).first()
        
    if not item: return

    reporter = item.reporter
    target_type = "FOUND" if item_type.upper() == "LOST" else "LOST"
    
    # Text-based matching
    potential_matches = db.query(LostItem).filter(
        LostItem.category == item.category,
        LostItem.type == target_type,
        LostItem.reporter != reporter,
        LostItem.status.in_(["Approved", "Verified", "Active", "IN CUSTODY"])
    ).all()

    for match in potential_matches:
        if item.item_name.lower() in match.item_name.lower() or match.item_name.lower() in item.item_name.lower():
            item.status = "Potential Match Found"
            if hasattr(item, 'matched_with'):
                item.matched_with = match.id
            if hasattr(match, 'matched_with'):
                if not match.matched_with:
                    match.status = "Potential Match Found"
                    match.matched_with = item.id
            
            db.add(Notification(recipient=reporter, message=f"System Match! A potential match for your '{item.item_name}' was found: '{match.item_name}'. Check your account page!"))
            db.add(Notification(recipient=match.reporter, message=f"System Match! Someone just reported an item that matches your '{match.item_name}'. Check your account page!"))

            await manager.send_personal_message({"type": "notification", "message": f"System Match! A potential match for your '{match.item_name}' was just reported! 🛡️"}, match.reporter)
            
            # Email Notification
            reporter_user = db.query(User).filter(User.username.ilike(match.reporter)).first()
            if not reporter_user:
                # Fallback: search by email prefix
                reporter_user = db.query(User).filter(User.email.ilike(f"{match.reporter}@%")).first()
            
            if reporter_user and reporter_user.email:
                try:
                    send_email_notification(
                        recipient_email=reporter_user.email, 
                        subject="Potential Match Found!",
                        message_content=f"Hello {reporter_user.first_name},\n\nGood news! A potential match for your reported item '{match.item_name}' has been found.\n\nPlease log in to TrackBox to check the details.\n\nBest regards,\nTrackBox Team"
                    )
                except: pass
    db.commit()

# ======================= SEARCH LOST ITEMS =======================
@app.get("/student/search", response_class=HTMLResponse)
def student_search(request: Request, token: Optional[str] = None, db: Session = Depends(get_db)):
    # STRICT authetication via cookie
    auth_token = request.cookies.get("access_token")
    if not auth_token:
         return RedirectResponse("/login")

    try:
        data = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
        if data["role"] != "student":
            return RedirectResponse("/login")
    except:
        return RedirectResponse("/login")

    # Fetch only items NOT reported by the current user to avoid redundancy, and NOT yet claimed/archived
    # Fetch only items NOT reported by the current user to avoid redundancy, and NOT yet claimed/archived/returned
    lost_items = db.query(LostItem).filter(
        LostItem.reporter != data["sub"], 
        LostItem.is_archived == False,
        LostItem.status.in_(["Approved", "Verified", "Active"]), # Only verified items
        ~LostItem.status.contains("RETURNED"),
        ~LostItem.status.contains("CLAIMED")
    ).order_by(LostItem.created_at.desc()).limit(10).all()
    found_items = db.query(FoundItem).filter(
        FoundItem.reporter != data["sub"], 
        FoundItem.is_archived == False, 
        FoundItem.is_claimed == False,
        FoundItem.status.in_(["Approved", "Verified", "Active", "IN CUSTODY"]) # Only verified items
    ).order_by(FoundItem.created_at.desc()).limit(10).all()

    all_recent = []
    for i in lost_items:
        all_recent.append({
            "id": i.id,
            "item_name": i.item_name,
            "category": i.category,
            "location": i.last_seen,
            "image": i.image,
            "type": i.type.lower() if i.type else "lost",
            "status": i.status,
            "is_claimed": False,
            "created_at": i.created_at
        })
    for i in found_items:
        all_recent.append({
            "id": i.id,
            "item_name": i.item_name,
            "category": i.category,
            "location": i.found_in,
            "image": i.image,
            "type": "found",
            "status": i.status,
            "is_claimed": i.is_claimed,
            "created_at": i.created_at
        })
    
    # Sort combined recent items
    all_recent.sort(key=lambda x: x["created_at"], reverse=True)

    response = templates.TemplateResponse("search_lost.html", {
        "request": request,
        "token": auth_token,
        "lost_items": all_recent, # Rename or keep as is for template compatibility
        "user": get_user(db, data["sub"])
    })
    # Prevent browser from caching this page
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    return response


@app.post("/student/search")
async def search_items_post(
    request: Request,
    token: str = Form(...),
    query: str = Form(""),
    image: UploadFile = File(None),
    db: Session = Depends(get_db)
):
    try:
        data = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        if data["role"] != "student":
            raise Exception()
    except:
        return JSONResponse({"error": "Invalid token"}, status_code=401)

    text_results = []
    if query:
        lost_text_results = db.query(LostItem).filter(
            (LostItem.reporter != data["sub"]),
            (LostItem.is_archived == False),
            (LostItem.status.in_(["Approved", "Verified", "Active"])),
            (~LostItem.status.contains("RETURNED")),
            (~LostItem.status.contains("CLAIMED")),
            (LostItem.item_name.contains(query)) |
            (LostItem.description.contains(query))
        ).all()
        
        for item in lost_text_results:
            text_results.append({
                "id": item.id,
                "reporter": item.reporter,
                "item_name": item.item_name,
                "category": item.category,
                "description": item.description,
                "location": item.last_seen,
                "image": item.image,
                "type": item.type.lower() if item.type else "lost",
                "status": item.status,
                "is_claimed": False
            })

        # Also search in found items (excluding own)
        found_text_results = db.query(FoundItem).filter(
            (FoundItem.reporter != data["sub"]),
            (FoundItem.description.contains(query)) |
            (FoundItem.item_name.contains(query)),
            (FoundItem.is_archived == False),
            (FoundItem.is_claimed == False),
            (FoundItem.status.in_(["Approved", "Verified", "Active", "IN CUSTODY"]))
        ).all()

        # Add found items to text results with a type indicator
        for item in found_text_results:
            text_results.append({
                "id": item.id,
                "reporter": item.reporter,
                "item_name": item.item_name,
                "category": item.category,
                "description": item.description,
                "location": item.found_in,
                "image": item.image,
                "type": "found",
                "status": item.status,
                "is_claimed": item.is_claimed
            })

    similar_items = []
    if image and image.filename:
        # Ensure directory exists
        os.makedirs("temp_uploads", exist_ok=True)
        # Save the uploaded image temporarily
        temp_image_path = os.path.join("temp_uploads", image.filename)
        with open(temp_image_path, "wb") as buffer:
            shutil.copyfileobj(image.file, buffer)

        # Compare with all lost items that have images (excluding own, excluding archived, excluding returned)
        lost_items_with_images = db.query(LostItem).filter(
            LostItem.image.isnot(None), 
            LostItem.reporter != data["sub"], 
            LostItem.is_archived == False,
            (LostItem.status.in_(["Approved", "Verified", "Active"])),
            (~LostItem.status.contains("RETURNED")),
            (~LostItem.status.contains("CLAIMED"))
        ).all()
        for item in lost_items_with_images:
            item_image_path = os.path.join("uploads", item.image)
            if os.path.exists(item_image_path):
                similarity = calculate_image_similarity(temp_image_path, item_image_path)
                if similarity > 0.5:  # Threshold for similarity
                    similar_items.append({
                        "id": item.id,
                        "reporter": item.reporter,
                        "item_name": item.item_name,
                        "category": item.category,
                        "description": item.description,
                        "last_seen": item.last_seen,
                        "location": item.last_seen,
                        "image": item.image,
                        "similarity": similarity,
                        "type": item.type.lower() if item.type else "lost",
                        "status": item.status,
                        "is_claimed": False
                    })

        # Compare with all found items that have images (excluding own, excluding archived, excluding claimed)
        found_items_with_images = db.query(FoundItem).filter(
            FoundItem.image.isnot(None), 
            FoundItem.reporter != data["sub"], 
            FoundItem.is_archived == False,
            (FoundItem.is_claimed == False),
            (FoundItem.status.in_(["Approved", "Verified", "Active", "IN CUSTODY"]))
        ).all()
        for item in found_items_with_images:
            item_image_path = os.path.join("uploads", item.image)
            if os.path.exists(item_image_path):
                similarity = calculate_image_similarity(temp_image_path, item_image_path)
                if similarity > 0.5:  # Threshold for similarity
                    similar_items.append({
                        "id": item.id,
                        "reporter": item.reporter,
                        "item_name": item.item_name,
                        "category": item.category,
                        "description": item.description,
                        "found_in": item.found_in,
                        "location": item.found_in,
                        "image": item.image,
                        "similarity": similarity,
                        "type": "found",
                        "status": item.status,
                        "is_claimed": item.is_claimed
                    })

        # Remove the temporary image
        os.remove(temp_image_path)

    # Combine text search and image search results
    combined_results = text_results + similar_items

    # Remove duplicates
    unique_results = []
    seen_ids = set()
    for item in combined_results:
        item_id = item.get("id", getattr(item, 'id', None))
        if item_id not in seen_ids:
            seen_ids.add(item_id)
            unique_results.append(item)

    # Fetch user for the template context
    user = get_user(db, data["sub"])
    # Fetch original list excluding own reports
    all_lost_items = db.query(LostItem).filter(LostItem.reporter != data["sub"]).order_by(LostItem.created_at.desc()).all()

    return templates.TemplateResponse("search_lost.html", {
        "request": request,
        "token": token,
        "user": data,
        "results": unique_results,
        "query": query,
        "lost_items": all_lost_items, # Kept all_lost_items as lost_items was undefined
        "searched": True if query or (image and image.filename) else False,
        "search_type": "image" if (image and image.filename) else "text"
    })
        


# ======================= CLAIM ITEM =======================
@app.get("/claim_qr", response_class=HTMLResponse)
def claim_qr(request: Request):
    return templates.TemplateResponse("claim_qr.html", {"request": request})

@app.post("/admin/move-to-found")
async def move_to_found(
    request: Request,
    token: str = Form(...),
    item_id: int = Form(...),
    found_in: Optional[str] = Form("Verified by Admin"),
    image: Optional[UploadFile] = File(None),
    db: Session = Depends(get_db)
):
    try:
        verify_admin_token(token)
        
        item = db.query(LostItem).filter(LostItem.id == item_id).first()
        if not item:
            return JSONResponse({"success": False, "message": "Item not found"})

        # Check if image is provided
        filename = item.image
        if image and image.filename:
            timestamp = int(get_ph_time().timestamp())
            filename = f"{timestamp}_{image.filename}"
            file_path = os.path.join(UPLOAD_FOLDER, filename)
            with open(file_path, "wb") as f:
                shutil.copyfileobj(image.file, f)

        new_found = FoundItem(
            reporter=item.reporter,
            category=item.category,
            item_name=item.item_name,
            description=item.description,
            found_in=found_in,
            image=filename,
            created_at=get_ph_time()
        )
        db.add(new_found)

        # Notify the reporter
        notification = Notification(
            recipient=item.reporter,
            message=f"Update: Your item '{item.item_name}' has been verified as FOUND and moved to the Found Items database."
        )
        db.add(notification)
        
        db.delete(item)
        db.commit()
        return JSONResponse({"success": True, "message": "Item verified and moved to Found Items successfully"})
    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": f"Database error: {str(e)}"}, status_code=500)

@app.post("/admin/mark-claimed")
async def mark_item_claimed(
    request: Request,
    token: str = Form(None),
    item_id: int = Form(...),
    db: Session = Depends(get_db)
):
    try:
        # Accept token from form body OR fallback to session cookie
        auth_token = token or request.cookies.get("access_token")
        if not auth_token:
            return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
        verify_admin_token(auth_token)
    except Exception:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)

    item = db.query(FoundItem).filter(FoundItem.id == item_id).first()
    if not item:
        # Fallback to LostItem if not in FoundItem
        item = db.query(LostItem).filter(LostItem.id == item_id).first()
        if item:
            # For LostItem, we mark it by moving it to FoundItem and immediately claiming it, 
            # or just deleting it. Since we want a 'Returned' record, we move it to FoundItem first.
            new_found = FoundItem(
                reporter=item.reporter,
                category=item.category,
                item_name=item.item_name,
                description=item.description,
                found_in="Directly Returned via Scanner",
                image=item.image,
                is_claimed=True,
                status="RETURNED TO OWNER",
                handed_over_by="Admin Office",
                handover_method="Admin Scanner",
                claimed_at=get_ph_time(),
                created_at=get_ph_time()
            )
            db.add(new_found)
            db.delete(item)
            db.commit()
            db.refresh(new_found)
            item = new_found # Use the new record for notification
    else:
        item.is_claimed = True
        item.status = "RETURNED TO OWNER"
        item.handed_over_by = "Admin Office"
        item.handover_method = "Admin Scanner"
        item.claimed_at = get_ph_time()
        db.commit()  # Commit the status change immediately
    
    # Notify the reporter
    notification = Notification(
        recipient=item.reporter,
        message=f"Success! Your item '{item.item_name}' has been officially marked as RETURNED to you via QR Scan!"
    )
    db.add(notification)
    
    # Award Points to Reporter (if they are a student)
    reward_points = 100
    reporter_user = db.query(User).filter(User.username == item.reporter).first()
    if reporter_user:
        reporter_user.points += reward_points
        db.add(Notification(
            recipient=item.reporter,
            message=f"You earned {reward_points} Good Samaritan points for returning an item! Keep it up! 🏆"
        ))
    
    db.commit()
    return JSONResponse({"success": True, "message": "Item marked as RETURNED to you via QR Scan!"})

@app.post("/notifications/mark-read")
async def mark_notifications_read(request: Request, db: Session = Depends(get_db)):
    auth_token = request.cookies.get("access_token")
    if not auth_token:
        return JSONResponse({"status": "error"}, status_code=401)
    
    try:
        user_data = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
        username = user_data["sub"]
        
        # Update all unread notifications for this user
        db.query(Notification).filter(
            Notification.recipient == username,
            Notification.is_read == False
        ).update({"is_read": True})
        
        db.commit()
        return {"success": True}
    except Exception as e:
        return JSONResponse({"status": "error", "detail": str(e)}, status_code=500)

@app.post("/student/preview-claim")
async def preview_claim(
    request: Request,
    qr_data: str = Form(...),
    db: Session = Depends(get_db)
):
    # Authenticate (Basic check)
    if not request.cookies.get("access_token"):
         return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
         
    # Parse QR Data
    if not qr_data.startswith("CLAIM ITEM"):
        return JSONResponse({"success": False, "message": "Invalid QR Code format. Please scan a valid Owner Claim QR."})

    try:
        parts = qr_data.replace("CLAIM ITEM ", "").split(" BY ")
        if len(parts) != 2:
             return JSONResponse({"success": False, "message": "Corrupted QR Data."})
             
        lost_item_id = int(parts[0])
        owner_username = parts[1]
        
        # Fetch Details
        lost_item = db.query(LostItem).filter(LostItem.id == lost_item_id).first()
        if not lost_item:
             return JSONResponse({"success": False, "message": f"The item referenced in this QR (ID #{lost_item_id}) no longer exists."})
             
        # Return Preview Data
        return {
            "success": True,
            "data": {
                "item_name": lost_item.item_name,
                "category": lost_item.category,
                "reporter": lost_item.reporter,
                "image": lost_item.image,
                "created_at": lost_item.created_at.strftime("%Y-%m-%d"),
                "status": lost_item.status
            }
        }
        
    except Exception as e:
        return JSONResponse({"success": False, "message": f"Error processing QR: {str(e)}"})

@app.post("/student/peer-verify")
async def peer_verify(
    request: Request,
    token: Optional[str] = Form(None),
    qr_data: str = Form(...),
    found_item_id: int = Form(...),
    db: Session = Depends(get_db)
):
    try:
        # Get authenticated user (the finder)
        auth_token = request.cookies.get("access_token") or token
        if not auth_token:
            return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
            
        payload = jwt.decode(auth_token, SECRET_KEY, algorithms=[ALGORITHM])
        finder_username = payload.get("sub")
    except Exception as e:
        return JSONResponse({"success": False, "message": "Authentication failed"}, status_code=401)

    try:
        # Parse Owner's QR (Supports legacy "CLAIM ITEM..." and new compact "TBX|..." format)
        owner_item_id = None
        owner_username = None

        if qr_data.startswith("CLAIM ITEM"):
            try:
                parts = qr_data.split()
                owner_item_id = int(parts[2])
                owner_username = parts[4]
            except:
                return JSONResponse({"success": False, "message": "Malformed legacy QR data."})
        elif qr_data.startswith("TBX|"):
            try:
                parts = qr_data.split("|")
                owner_item_id = int(parts[1])
                owner_username = parts[2]
            except:
                return JSONResponse({"success": False, "message": "Malformed compact QR data."})
        else:
            return JSONResponse({"success": False, "message": "Unrecognized QR code format."})

        # Get the finder's found report - check both tables
        # First try FoundItem table
        found_report = db.query(FoundItem).filter(FoundItem.id == found_item_id).first()
        
        # If not found, try LostItem table where type='FOUND'
        found_report_from_lost = None
        if not found_report:
            found_report_from_lost = db.query(LostItem).filter(
                LostItem.id == found_item_id,
                LostItem.type == "FOUND"
            ).first()
        
        # Get the owner's lost report
        lost_report = db.query(LostItem).filter(LostItem.id == owner_item_id).first()

        # IMMEDIATE PROCESSING - No validation, just do it
        if found_report:
            # Item is in FoundItem table
            found_report.is_claimed = True
            found_report.status = "RETURNED (PEER TO PEER)"
            found_report.handed_over_by = finder_username
            found_report.handed_over_to = owner_username
            found_report.handover_method = "Peer-to-Peer"
            found_report.claimed_at = get_ph_time()
            # Don't archive - keep visible for admin tracking
            found_report.description += f" [Returned to owner: {owner_username}]"
        elif found_report_from_lost:
            # Item is in LostItem table with type='FOUND'
            found_report_from_lost.status = "RETURNED (PEER TO PEER)"
            found_report_from_lost.handed_over_by = finder_username
            found_report_from_lost.handed_over_to = owner_username
            found_report_from_lost.handover_method = "Peer-to-Peer"
            found_report_from_lost.claimed_at = get_ph_time()
            # Don't archive FOUND items - keep visible for admin
            found_report_from_lost.description += f" [Returned to owner: {owner_username}]"
        
        if lost_report and lost_report.type == "LOST":
            # Only archive actual LOST reports
            lost_report.status = "RETURNED (PEER TO PEER)"
            lost_report.handed_over_by = finder_username
            lost_report.handover_method = "Peer-to-Peer"
            lost_report.claimed_at = get_ph_time()
            lost_report.is_archived = True

        # Award Points to Finder
        finder_user = db.query(User).filter(User.username == finder_username).first()
        if finder_user:
            finder_user.points += 150

        # Notify Owner
        db.add(Notification(
            recipient=owner_username,
            message=f"Success! {finder_username} has handed over your item to you directly!"
        ))

        # Notify Admin
        item_name = found_report.item_name if found_report else lost_report.item_name if lost_report else "Unknown Item"
        db.add(Notification(
            recipient="admin",
            message=f"Peer Handover: Item '{item_name}' was successfully returned to {owner_username} by {finder_username}."
        ))
        
        db.commit()
        return JSONResponse({"success": True, "message": "Handover successful! Points awarded."})

    except Exception as e:
        db.rollback()
        return JSONResponse({"success": False, "message": f"Handover processing error: {str(e)}"})

# ======================= MESSAGING =======================
@app.post("/messages/send")
async def send_message(
    request: Request,
    receiver: str = Form(...),
    content: str = Form(...),
    item_id: int = Form(...),
    db: Session = Depends(get_db)
):
    user = get_current_user(request, db)
    if not user:
        return JSONResponse({"success": False, "message": "Unauthorized"}, status_code=401)
    
    msg = Message(
        sender=user.username,
        receiver=receiver,
        content=content,
        item_id=item_id
    )
    db.add(msg)
    
    # Notify Receiver
    db.add(Notification(
        recipient=receiver,
        message=f"New message from {user.username} regarding item #{item_id}: '{content[:20]}...'"
    ))
    db.commit()

    # Real-time Broadcast
    await manager.send_personal_message({
        "type": "message",
        "sender": user.username,
        "content": content,
        "item_id": item_id
    }, receiver)

    return {"success": True}

@app.get("/messages/{item_id}")
def get_messages(item_id: int, request: Request, db: Session = Depends(get_db)):
    user = get_current_user(request, db)
    if not user:
        return []
    
    messages = db.query(Message).filter(
        Message.item_id == item_id,
        (Message.sender == user.username) | (Message.receiver == user.username)
    ).order_by(Message.created_at.asc()).all()
    
    return [{
        "id": m.id, 
        "sender": m.sender, 
        "receiver": m.receiver, 
        "content": m.content, 
        "created_at": m.created_at.isoformat() if m.created_at else None
    } for m in messages]

# ======================= ADMIN ANALYTICS =======================
@app.get("/admin/stats")
def get_stats(building: Optional[str] = None, db: Session = Depends(get_db)):
    # Base queries
    lost_q = db.query(LostItem).filter(LostItem.is_archived == False)
    found_q = db.query(FoundItem).filter(FoundItem.is_claimed == False)
    
    # Returned items query (lost items archived as returned + claimed found items)
    lost_returned_q = db.query(LostItem).filter(LostItem.is_archived == True, LostItem.status != "Rejected")
    found_returned_q = db.query(FoundItem).filter(FoundItem.is_claimed == True)
    
    # Optional building filter (US-03)
    if building and building != "All Campus" and building != "[object Event]":
        lost_q = lost_q.filter(LostItem.building == building)
        found_q = found_q.filter(FoundItem.building == building)
        lost_returned_q = lost_returned_q.filter(LostItem.building == building)
        found_returned_q = found_returned_q.filter(FoundItem.building == building)

    lost_count = lost_q.count()
    found_count = found_q.count()
    returned_count = lost_returned_q.count() + found_returned_q.count()

    # Category distribution for the selected building/campus (combining Lost and Found)
    categories = {}
    
    # Active Lost Items categories
    for item in lost_q.all():
        if item.category:
            categories[item.category] = categories.get(item.category, 0) + 1
            
    # Unclaimed Found Items categories
    for item in found_q.all():
        if item.category:
            categories[item.category] = categories.get(item.category, 0) + 1

    return {
        "lost": lost_count,
        "found": found_count,
        "returned": returned_count,
        "categories": categories,
        "building": building or "All Campus"
    }


# ======================= BUILDING MANAGEMENT (Dynamic) =======================

@app.get("/api/buildings")
def get_buildings_api(db: Session = Depends(get_db)):
    buildings = db.query(Building).all()
    if not buildings:
        # Seed if empty
        defaults = ["Main Building", "Engineering Building", "Computer Studies Building", "Security Post"]
        for b in defaults:
            db.add(Building(name=b))
        db.commit()
        buildings = db.query(Building).all()
    return [{"id": b.id, "name": b.name} for b in buildings]

@app.post("/admin/buildings/add")
def add_building(token: str = Form(...), name: str = Form(...), db: Session = Depends(get_db)):
    try:
        verify_admin_token(token)
        if db.query(Building).filter(Building.name == name).first():
            return JSONResponse({"success": False, "message": "Building already exists"})
        db.add(Building(name=name))
        db.commit()
        return {"success": True, "message": "Building added"}
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

@app.post("/admin/buildings/delete")
def delete_building(token: str = Form(...), b_id: int = Form(...), db: Session = Depends(get_db)):
    try:
        verify_admin_token(token)
        b = db.query(Building).filter(Building.id == b_id).first()
        if b:
            db.delete(b)
            db.commit()
        return {"success": True, "message": "Building deleted"}
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})
# ======================= REPORTS & EXPORT (US-05) =======================
import csv
import io
import openpyxl
import zipfile
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse, FileResponse

@app.get("/admin/export/report")
def export_monthly_report(
    token: str, 
    month: int = None, 
    year: int = None, 
    building: Optional[str] = None,
    db: Session = Depends(get_db)
):
    try:
        verify_admin_token(token)
        
        # Base queries
        found_q = db.query(FoundItem)
        lost_q = db.query(LostItem)
        
        if building and building != "All Campus":
            found_q = found_q.filter(FoundItem.building == building)
            lost_q = lost_q.filter(LostItem.building == building)
            
        if month and year:
            # Filter by month and year (SQLite logic using strftime)
            m_str = f"{month:02d}"
            y_str = str(year)
            
            found_q = found_q.filter(
                func.strftime('%m', FoundItem.created_at) == m_str,
                func.strftime('%Y', FoundItem.created_at) == y_str
            )
            lost_q = lost_q.filter(
                func.strftime('%m', LostItem.created_at) == m_str,
                func.strftime('%Y', LostItem.created_at) == y_str
            )

        found_items = found_q.all()
        lost_items = lost_q.all()

        # Use OpenPyXL for professional Excel export with AutoFit
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TrackBox Report"
        
        # Styles
        title_font = Font(name='Arial', size=14, bold=True, color="800000")
        header_font = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        
        # --- LOGO SECTION ---
        # Adjust row heights for logos (Row 1-5 for header)
        ws.row_dimensions[1].height = 60 
        
        # Add EVSU Logo (Left)
        try:
            evsu_img = XLImage('EVSU_logo.png')
            evsu_img.width, evsu_img.height = 70, 70
            ws.add_image(evsu_img, 'A1')
        except: pass

        # Add Bagong Pilipinas Logo (Right) - Using user's filename typo 'bagon_pilipinas.jpg'
        try:
            bp_img = XLImage('bagon_pilipinas.jpg')
            bp_img.width, bp_img.height = 70, 70
            ws.add_image(bp_img, 'H1')
        except: pass

        # Report Title
        ws.merge_cells('B1:G1')
        ws['B1'] = "TRACKBOX MONTHLY REPORT"
        ws['B1'].font = title_font
        ws['B1'].alignment = center_align
        
        ws.append([]) # Row 2
        ws.append([f"Building: {building or 'All Campus'}", "", f"Period: {month}/{year if month else 'All Time'}"]) # Row 3
        ws.append([]) # Row 4
        
        # Borders
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Found Items Section
        curr_row = ws.max_row + 1
        ws.append(["FOUND ITEMS (OFFICIAL)"])
        ws.merge_cells(f'A{curr_row}:H{curr_row}')
        ws[f'A{curr_row}'].font = Font(bold=True, size=11, color="800000")
        
        headers_found = ["ID", "Item Name", "Category", "Building", "Location", "Status", "Date Found", "Reporter/Finder"]
        ws.append(headers_found)
        header_row = ws.max_row
        for cell in ws[header_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for i, item in enumerate(found_items):
            row = [item.id, item.item_name, item.category, item.building, item.found_in, item.status, item.created_at.strftime('%Y-%m-%d'), item.reporter]
            ws.append(row)
            curr_row_idx = ws.max_row
            fill = PatternFill(start_color="FFF4F4", end_color="FFF4F4", fill_type="solid") if i % 2 == 1 else None
            for cell in ws[curr_row_idx]:
                if fill: cell.fill = fill
                cell.border = thin_border
            
        ws.append([])
        
        # Lost Items Section
        curr_row = ws.max_row + 1
        ws.append(["LOST & FOUND REPORTS (STUDENT)"])
        ws.merge_cells(f'A{curr_row}:I{curr_row}')
        ws[f'A{curr_row}'].font = Font(bold=True, size=11, color="800000")
        
        headers_lost = ["ID", "Type", "Item Name", "Category", "Building", "Location", "Status", "Date Reported", "Reporter"]
        ws.append(headers_lost)
        header_row_lost = ws.max_row
        for cell in ws[header_row_lost]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for i, item in enumerate(lost_items):
            row = [item.id, item.type, item.item_name, item.category, item.building, item.last_seen, item.status, item.created_at.strftime('%Y-%m-%d'), item.reporter]
            ws.append(row)
            curr_row_idx = ws.max_row
            fill = PatternFill(start_color="FFF4F4", end_color="FFF4F4", fill_type="solid") if i % 2 == 1 else None
            for cell in ws[curr_row_idx]:
                if fill: cell.fill = fill
                cell.border = thin_border

        # --- AUTOFIT COLUMNS ---
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                # Skip merged cells for length calculation to avoid errors
                if hasattr(cell, 'value') and cell.value:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = (max_length + 4)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"TrackBox_Report_{building or 'All'}_{month or 'Total'}_{year or ''}.xlsx"
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.post("/api/confirm-match")
def confirm_match(item_id: int = Form(...), db: Session = Depends(get_db)):
    try:
        item = db.query(LostItem).filter(LostItem.id == item_id).first()
        if not item:
            return JSONResponse({"success": False, "message": "Item not found"})
        
        item.is_owner_verified = True
        item.verified_at = get_ph_time()
        
        # If it's a mutual match, verify the other one too
        if item.matched_with:
            other = db.query(LostItem).filter(LostItem.id == item.matched_with).first()
            if other:
                other.is_owner_verified = True
                other.verified_at = get_ph_time()
        
        db.commit()
        return {"success": True, "message": "Item confirmed! You can now chat with the finder."}
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)})

# ======================= DATA BACKUP (US-06) =======================
@app.get("/admin/backup/download")
def download_backup(token: str, db: Session = Depends(get_db)):
    try:
        verify_admin_token(token)
        
        from sqlalchemy.inspection import inspect
        import zipfile
        import io
        from datetime import datetime
        
        models = [Building, User, LostItem, FoundItem, Notification, Message]
        sql_lines = []
        sql_lines.append("-- TrackBox Database Backup (MySQL Compatible)")
        sql_lines.append(f"-- Generated At: {get_ph_time().strftime('%Y-%m-%d %H:%M:%S')} PH Time")
        sql_lines.append("SET FOREIGN_KEY_CHECKS = 0;\n")
        
        for model in models:
            table_name = model.__tablename__
            sql_lines.append(f"-- ------------------------------------------------------")
            sql_lines.append(f"-- Table structure and data for table `{table_name}`")
            sql_lines.append(f"-- ------------------------------------------------------")
            sql_lines.append(f"DROP TABLE IF EXISTS `{table_name}`;")
            
            columns = inspect(model).mapper.columns
            create_parts = []
            for col in columns:
                col_name = col.key
                col_type = col.type
                
                # Map SQLAlchemy type to MySQL type
                mysql_type = "VARCHAR(255)"
                if str(col_type).startswith("INTEGER"):
                    mysql_type = "INT"
                elif str(col_type).startswith("BOOLEAN"):
                    mysql_type = "TINYINT(1)"
                elif str(col_type).startswith("DATETIME"):
                    mysql_type = "DATETIME"
                elif str(col_type).startswith("TEXT"):
                    mysql_type = "TEXT"
                
                nullable = "NULL" if col.nullable else "NOT NULL"
                if col.primary_key:
                    nullable += " AUTO_INCREMENT"
                
                create_parts.append(f"  `{col_name}` {mysql_type} {nullable}")
                
            pk_cols = [c.key for c in columns if c.primary_key]
            if pk_cols:
                pk_str = ", ".join([f"`{c}`" for c in pk_cols])
                create_parts.append(f"  PRIMARY KEY ({pk_str})")
                
            create_sql = f"CREATE TABLE `{table_name}` (\n" + ",\n".join(create_parts) + "\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;\n"
            sql_lines.append(create_sql)
            
            # Fetch rows
            rows = db.query(model).all()
            if rows:
                col_names = [col.key for col in columns]
                col_str = ", ".join([f"`{c}`" for c in col_names])
                
                for row in rows:
                    vals = []
                    for col in col_names:
                        val = getattr(row, col)
                        if val is None:
                            vals.append("NULL")
                        elif isinstance(val, bool):
                            vals.append("1" if val else "0")
                        elif isinstance(val, (int, float)):
                            vals.append(str(val))
                        elif isinstance(val, datetime):
                            vals.append(f"'{val.strftime('%Y-%m-%d %H:%M:%S')}'")
                        else:
                            escaped = str(val).replace("\\", "\\\\").replace("'", "\\'")
                            vals.append(f"'{escaped}'")
                    
                    val_str = ", ".join(vals)
                    sql_lines.append(f"INSERT INTO `{table_name}` ({col_str}) VALUES ({val_str});")
            sql_lines.append("")
            
        sql_lines.append("SET FOREIGN_KEY_CHECKS = 1;")
        sql_content = "\n".join(sql_lines)
        
        # Create ZIP archive in memory
        timestamp = get_ph_time().strftime("%Y%m%d_%H%M%S")
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            zip_file.writestr(f"trackbox_{timestamp}.sql", sql_content)
        
        zip_buffer.seek(0)
        filename = f"TrackBox_MySQL_Backup_{timestamp}.zip"
        
        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)

@app.websocket("/ws/{username}")
async def websocket_endpoint(websocket: WebSocket, username: str):
    await manager.connect(username, websocket)
    try:
        while True:
            data = await websocket.receive_json()
            receiver = data.get("receiver")
            chat_id = data.get("chat_id")
            text = data.get("text")
            
            # Save to Database if valid message
            if receiver and chat_id and text:
                db = SessionLocal()
                try:
                    msg = Message(
                        sender=username,
                        receiver=receiver,
                        content=text,
                        item_id=int(chat_id)
                    )
                    db.add(msg)
                    db.add(Notification(
                        recipient=receiver,
                        message=f"New chat message from {username}: '{text[:30]}...'"
                    ))
                    db.commit()
                except Exception as e:
                    print(f"Error saving websocket chat message: {e}")
                    db.rollback()
                finally:
                    db.close()
            
            if receiver:
                await manager.send_personal_message(data, receiver)
    except WebSocketDisconnect:
        manager.disconnect(username)


@app.get("/download_qr")
async def download_qr(data: str):
    """
    Proxies the QR code from the external API to allow browser downloads (avoids CORS issues).
    """
    async with httpx.AsyncClient() as client:
        qr_url = f"https://api.qrserver.com/v1/create-qr-code/?size=250x250&data={data}&ecc=M"
        response = await client.get(qr_url)
        if response.status_code != 200:
            raise HTTPException(status_code=500, detail="Failed to fetch QR code")
        
        return StreamingResponse(
            io.BytesIO(response.content),
            media_type="image/png",
            headers={"Content-Disposition": "attachment; filename=trackbox-qr.png"}
        )

@app.post("/user/mark-as-found/{item_id}")
async def mark_as_found_user(
    item_id: int,
    request: Request,
    db: Session = Depends(get_db)
):
    user = get_current_user(request, db)
    if not user:
        return RedirectResponse("/login")
    
    item = db.query(LostItem).filter(LostItem.id == item_id, LostItem.reporter == user.username).first()
    if not item:
        return RedirectResponse("/user_account")
    
    if item.type == "LOST":
        item.status = "Found by Owner"
        # We could also move it to FoundItem table, but let's keep it in LostItem for history
        # Or Just change type to FOUND
        item.type = "FOUND"
        
        db.add(Notification(
            recipient=user.username,
            message=f"You have marked '{item.item_name}' as found. Great!"
        ))
        db.commit()
        
    return RedirectResponse("/user_account", status_code=303)
